import os
import pandas as pd
import shutil
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from typing import List, Dict, Any
from datetime import datetime
from .models import ExcelItem

class ExcelService:
    """Excel文件服务类"""
    
    def __init__(self, base_dir: str = "data/excel_files"):
        self.base_dir = base_dir
        # 确保目录存在
        os.makedirs(base_dir, exist_ok=True)
        # 备份目录
        self.backup_dir = os.path.join(base_dir, "backups")
        os.makedirs(self.backup_dir, exist_ok=True)
    
    def list_excel_files(self) -> List[str]:
        """列出所有Excel文件"""
        try:
            return [
                f for f in os.listdir(self.base_dir)
                if f.endswith(".xlsx") and f != "temp.xlsx"
            ]
        except FileNotFoundError:
            return []
    
    def read_excel(self, file_name: str) -> Dict[str, Any]:
        """读取Excel文件数据"""
        path = os.path.join(self.base_dir, file_name)
        
        if not os.path.exists(path):
            raise FileNotFoundError(f"文件不存在: {file_name}")
        
        df = pd.read_excel(path)
        
        # 确保必要的列存在
        if "数量" in df.columns and "价格" in df.columns:
            df["总价"] = df["数量"] * df["价格"]
            total = df["总价"].sum()
        else:
            total = 0
        
        # 转换为字典列表
        records = []
        for _, row in df.iterrows():
            item = {}
            for col in df.columns:
                item[col] = row[col] if pd.notna(row[col]) else None
            records.append(item)
        
        return {
            "records": records,
            "total": float(total)
        }
    
    def save_excel(self, file_name: str, records: List[Dict[str, Any]]) -> bool:
        """保存数据到Excel文件"""
        path = os.path.join(self.base_dir, file_name)
        temp_path = os.path.join(self.base_dir, "temp.xlsx")
        # 备份放到备份目录，保留原始文件名并添加 .bak 后缀
        backup_path = os.path.join(self.backup_dir, f"{file_name}.bak")
        
        # 验证数据
        self._validate_records(records)
        
        # 添加序号
        records_with_index = self._add_index(records)
        
        # 转换为DataFrame
        df = pd.DataFrame(records_with_index)
        
        # 计算总价
        if "数量" in df.columns and "价格" in df.columns:
            df["总价"] = df["数量"] * df["价格"]
        
        # 备份原文件
        if os.path.exists(path):
            shutil.copy(path, backup_path)
        
        # 写入临时文件
        df.to_excel(temp_path, index=False)
        
        # 替换原文件
        os.replace(temp_path, path)
        
        # 处理图片
        self._insert_images(file_name)
        
        return True
    
    def undo(self, file_name: str) -> bool:
        """撤回上次保存"""
        path = os.path.join(self.base_dir, file_name)
        backup_path = os.path.join(self.backup_dir, f"{file_name}.bak")

        if os.path.exists(backup_path):
            os.replace(backup_path, path)
            return True
        return False
    
    def _insert_images(self, file_name: str):
        """将图片插入Excel"""
        path = os.path.join(self.base_dir, file_name)
        
        if not os.path.exists(path):
            return
        
        try:
            wb = load_workbook(path)
            ws = wb.active
            
            # 假设图片路径在第10列（J列）
            for row in range(2, ws.max_row + 1):
                cell = ws[f"J{row}"]
                image_path = cell.value
                
                # 如果是相对路径，优先在 data/pictures 下查找
                candidate = image_path
                if image_path and not os.path.isabs(image_path):
                    candidate = os.path.join("data", "pictures", image_path)

                if image_path and os.path.exists(candidate):
                    try:
                        img = XLImage(candidate)
                        img.width = 80
                        img.height = 80
                        ws.add_image(img, f"J{row}")
                    except Exception as e:
                        print(f"插入图片失败: {e}")
            
            wb.save(path)
        except Exception as e:
            print(f"处理图片时出错: {e}")
    
    def _validate_records(self, records: List[Dict[str, Any]]):
        """验证数据记录"""
        for r in records:
            if not r.get("内容"):
                raise ValueError("内容不能为空")
            if r.get("数量", 0) <= 0:
                raise ValueError("数量必须大于0")
            if r.get("价格", -1) < 0:
                raise ValueError("价格不能为负")
    
    def _add_index(self, records: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """添加序号"""
        for i, r in enumerate(records, start=1):
            r["序号"] = i
        return records
    
    def get_file_info(self, file_name: str) -> Dict[str, Any]:
        """获取文件信息"""
        path = os.path.join(self.base_dir, file_name)
        
        if not os.path.exists(path):
            raise FileNotFoundError(f"文件不存在: {file_name}")
        
        stat = os.stat(path)
        return {
            "name": file_name,
            "path": path,
            "size": stat.st_size,
            "modified": datetime.fromtimestamp(stat.st_mtime)
        }


# 创建全局服务实例
excel_service = ExcelService()
