import os
import pandas as pd
import shutil
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

def list_excel_files():
    return [
        f for f in os.listdir("data/excel_files")
        if f.endswith(".xlsx")
    ]

def read_excel(file_name):
    path = f"data/excel_files/{file_name}"
    df = pd.read_excel(path)

    df["总价"] = df["数量"] * df["价格"]
    total = df["总价"].sum()

    return {
        "records": df.to_dict(orient="records"),
        "total": float(total)
    }

def save_excel(file_name, records):
    base_path = f"data/excel_files/{file_name}"
    temp_path = f"data/excel_files/temp.xlsx"
    backup_path = f"data/excel_files/{file_name}.bak"
    df = pd.DataFrame(records)
    df["总价"] = df["数量"] * df["价格"]
    # 备份
    if os.path.exists(base_path):
        shutil.copy(base_path, backup_path)
    # 写入临时文件
    df.to_excel(temp_path, index=False)
    # 替换
    os.replace(temp_path, base_path)

def undo(file_name):
    base_path = f"data/excel_files/{file_name}"
    backup_path = f"{base_path}.bak"

    if os.path.exists(backup_path):
        os.replace(backup_path, base_path)
        return True
    return False

def insert_images(file_name):
    path = f"data/excel_files/{file_name}"
    wb = load_workbook(path)
    ws = wb.active
    for row in range(2, ws.max_row + 1):
        image_path = ws[f"J{row}"].value  # 假设图片路径

        if image_path and os.path.exists(image_path):
            img = XLImage(image_path)
            img.width = 80
            img.height = 80
            ws.add_image(img, f"J{row}")
    wb.save(path)

def validate(records):
    for r in records:
        if not r["内容"]:
            raise ValueError("内容不能为空")
        if r["数量"] <= 0:
            raise ValueError("数量必须大于0")
        if r["价格"] < 0:
            raise ValueError("价格不能为负")
        
def add_index(records):
    for i, r in enumerate(records, start=1):
        r["序号"] = i

def allow_delete(file_name):
    path = f"data/excel_files/{file_name}"
    df = pd.read_excel(path)
    return len(df) > 0

