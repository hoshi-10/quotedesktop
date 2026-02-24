#!/usr/bin/env python3
"""
诊断Excel数据读取问题
"""

import pandas as pd
import os
from openpyxl import load_workbook

def diagnose_excel_file(file_path):
    """诊断Excel文件的数据结构"""
    print(f"\n=== 诊断文件: {os.path.basename(file_path)} ===\n")
    
    # 1. 检查文件是否存在
    if not os.path.exists(file_path):
        print("❌ 文件不存在")
        return
    
    print("✅ 文件存在\n")
    
    # 2. 使用pandas读取
    try:
        df = pd.read_excel(file_path, dtype=str)
        print(f"✅ pandas读取成功")
        print(f"   行数: {len(df)}")
        print(f"   列数: {len(df.columns)}")
        print(f"   列名: {list(df.columns)}\n")
        
        # 显示前几行数据
        print("前5行数据:")
        print(df.head().to_string())
        print()
        
        # 检查每行的序号列
        print("序号列分析:")
        for idx, row in df.iterrows():
            seq_value = row.get("序号", None)
            print(f"   行{idx+1}: 序号='{seq_value}' (类型: {type(seq_value).__name__})")
            if idx >= 4:
                break
        print()
        
    except Exception as e:
        print(f"❌ pandas读取失败: {e}\n")
    
    # 3. 使用openpyxl读取
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        print(f"✅ openpyxl读取成功")
        print(f"   工作表名: {ws.title}")
        print(f"   最大行: {ws.max_row}")
        print(f"   最大列: {ws.max_column}\n")
        
        # 显示前几行
        print("前5行原始数据:")
        for row_idx in range(1, min(6, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                row_data.append(str(cell_value) if cell_value is not None else "")
            print(f"   行{row_idx}: {row_data}")
        print()
        
    except Exception as e:
        print(f"❌ openpyxl读取失败: {e}\n")
    
    # 4. 检查数据验证问题
    print("潜在问题分析:")
    
    # 检查列名
    expected_columns = ["序号", "内容", "材料", "规格尺寸", "数量", "价格", "总价", "项目图片", "经办人", "备注"]
    try:
        df = pd.read_excel(file_path, dtype=str)
        actual_columns = list(df.columns)
        
        missing_columns = set(expected_columns) - set(actual_columns)
        extra_columns = set(actual_columns) - set(expected_columns)
        
        if missing_columns:
            print(f"   ⚠️ 缺少列: {missing_columns}")
        if extra_columns:
            print(f"   ⚠️ 多余列: {extra_columns}")
        if not missing_columns and not extra_columns:
            print(f"   ✅ 列名完全匹配")
            
    except Exception as e:
        print(f"   ❌ 无法检查列名: {e}")
    
    # 检查数据类型
    try:
        df = pd.read_excel(file_path, dtype=str)
        print("\n数据类型检查:")
        
        for col in ["数量", "价格", "总价"]:
            if col in df.columns:
                non_null_values = df[col].dropna()
                if len(non_null_values) > 0:
                    print(f"   {col}:")
                    for idx, val in non_null_values.head(3).items():
                        try:
                            float_val = float(val)
                            print(f"      行{idx+1}: '{val}' -> {float_val} ✅")
                        except (ValueError, TypeError):
                            print(f"      行{idx+1}: '{val}' -> 转换失败 ❌")
                            
    except Exception as e:
        print(f"   ❌ 数据类型检查失败: {e}")
    
    print("\n=== 诊断完成 ===\n")

if __name__ == "__main__":
    excel_dir = "data/excel_files"
    
    # 检查目录中的所有Excel文件
    if os.path.exists(excel_dir):
        files = [f for f in os.listdir(excel_dir) if f.endswith(".xlsx") and f != "temp.xlsx"]
        
        for file_name in files:
            file_path = os.path.join(excel_dir, file_name)
            diagnose_excel_file(file_path)
    else:
        print(f"❌ 目录不存在: {excel_dir}")
